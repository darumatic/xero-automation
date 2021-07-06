#!/usr/bin/env python

import argparse
import calendar
import datetime
import json
import os
import random
import shutil
import sys
from pprint import pprint

import dateutil.parser
import pdfkit
import pystache
import pytz
from openpyxl import Workbook

from xero_client import XeroClient
from xero_email_sender import XeroEmailSender

DEFAULT_TIMEZONE = pytz.timezone("Australia/Sydney")

# This object is timezone aware and defaults to DEFAULT_TIMEZONE, it doesn't take into account $LOCAL_TIMEZONE
DARUMATIC_START_TIME = datetime.datetime(2017, 2, 1, tzinfo=DEFAULT_TIMEZONE).astimezone(pytz.utc)

class XeroReport:
    def __init__(self, args, command):
        if command not in set(["validate", "report-month", "report-po", "close"]):
            raise Exception(f"{command} isn't a valid command")

        self.command = command
        self.local_timezone = self._get_local_time_zone()
        # self.target_month_string is used for report generation
        self.target_month_string = self._get_time_sheet_project_name()

        self.MAX_DURATION = 12.0
        self.CURRENT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))


        # Times used by validate and close
        self.validate_close_start_time = None
        self.validate_close_end_time = None
        self.add_project_times(args.start_time, args.end_time)

        # Times used by timesheet generation (both monthly and by PO)
        self.timesheet_start_time = self._get_time_sheet_start_time(args.start_time)
        self.timesheet_end_time = self._get_time_sheet_end_time(args.end_time)


        self.client_id = args.client_id
        self.client_secret = args.client_secret
        self.tenant_id = args.tenant_id
        self.refresh_token = args.refresh_token
        self.cache_users = {}
        self.cache_tasks = {}
        self.duration_weeks = 2
        self.output = args.output
        self.skip_owners = args.skip_owners

        # For generating reports based on PO
        self.po = args.po if self.command == "report-po" else None

        self.suppress_email = args.suppress_email
        if not args.suppress_email and not self._email_env_var_present():
            # Override and suppress anyway if none of the required env vars are present
            self.suppress_email = True
            print("Email environmental variables missing, suppressing emails anyway")


        print("CI_PROJECT_ID=%s", os.getenv('CI_PROJECT_ID', None))
        self.xero_client = XeroClient(self.client_id, self.client_secret, self.tenant_id, self.refresh_token)
        self.DONT_VALIDATE_THESE_ITEMS = eval(os.environ.get('VALIDATION_EXCEPTIONS', '[]'))

        # Only load owner data if the flag is false
        if not self.skip_owners:
            #TODO: change this with the Xero Contacts data
            #example of environment variable
            #OWNERS = "{ 'Project A': 'Neil', 'Non chargeable tasks': 'Adrian' }"
            OWNERS_FILE = os.path.join(self.CURRENT_DIRECTORY, ".owners")
            if os.path.isfile(OWNERS_FILE):
                owners = open(OWNERS_FILE).read().strip()
            else:
                owners = os.environ['OWNERS']
            self.OWNERS = eval(owners)
            print(f"Owners: {self.OWNERS}")

        else:
            self.OWNERS = {}
            print("Skipping loading owner data")

        # Only load employees and holidays data for validation
        if self.command == "validate":
            # EMPLOYEES
            #EMPLOYEES_FILE = os.path.join(self.CURRENT_DIRECTORY, ".employees.json")
            #employees = open(EMPLOYEES_FILE).read().strip()
            #self.EMPLOYEES = eval(employees)
            # NSW Public Holidays
            NSW_HOLIDAYS_FILE = os.path.join(self.CURRENT_DIRECTORY, ".NSW_holidays.json")
            holidays = open(NSW_HOLIDAYS_FILE).read().strip()
            self.NSW_HOLIDAYS = eval(holidays)

        # Date sanity checks
        if (
            (self.command == "report-month" or self.command == "report-po")
            and (not (self.timesheet_start_time < self.timesheet_end_time))
        ):
            raise Exception("Timesheet end time is earlier than timesheet start time")

        if (
            (self.command == "validate" or self.command == "close")
            and (not (self.validate_close_start_time < self.validate_close_end_time))
        ):
            raise Exception("validate/close end time is earlier than validate/close start time")

        print(f"validate_close_start_time is {self.validate_close_start_time}")
        print(f"validate_close_end_time is {self.validate_close_end_time}")
        print(f"time_sheet_start_time is {self.timesheet_start_time}")
        print(f"time_sheet_end_time is {self.timesheet_end_time}")

    def add_project_times(self, start_time, end_time):
        """
        Sets the time ranges used by the validate and close function
        """
        # REVIEW Potential bug: time entries that are logged the second between 23:59:59 and 00:00:00 between the end and start of a new month may not be included.

        if self.command == "report-month" or self.command == "report-po":
            # Timesheet generation shouldn't use self.start_time and self.end_time
            return

        local_now = datetime.datetime.now(self.local_timezone)
        # self.filter is the current month
        self.filter = str(local_now)[0:4] + str(local_now)[5:7]
        print("Project Filter: {0}".format(self.filter))

        # For the following, the time (whether from start_time or self.filter) is in local time
        # We localise it with local timezone information and normalise it into UTC
        if start_time:
            # start_time has been passed in
            # Input is in local time
            start = datetime.datetime.strptime(start_time, "%Y-%m-%d")
            # Gives it timezone information and normalise to UTC
            self.validate_close_start_time = self.local_timezone.localize(start).astimezone(pytz.utc)

        if not self.validate_close_start_time:
            # self.start_time hasn't been set
            # Default to the start of the current month
            start = datetime.datetime.strptime(self.filter, "%Y%m")
            # Gives it timezone information and normalise to UTC
            self.validate_close_start_time = self.local_timezone.localize(start).astimezone(pytz.utc)

        if end_time:
            # end_time has been passed in
            end = datetime.datetime.strptime(end_time, "%Y-%m-%d")

            # Gives it timezone information
            end = self.local_timezone.localize(end)
            # Set it to the end of the day (in its local timezone)
            end = end.replace(hour=23, minute=59, second=59)
            # Normalise it to UTC
            self.validate_close_end_time = end.astimezone(pytz.utc)

        if not self.validate_close_end_time:
            # self.end_time hasn't been set
            month_last_day = calendar.monthrange(local_now.year, local_now.month)[1]
            end = datetime.datetime.strptime(self.filter, "%Y%m")

            # Gives it timezone information
            end = self.local_timezone.localize(end)
            # Defaults to the end of the current month
            end = end.replace(day=month_last_day, hour=23, minute=59, second=59)
            # Normalise it to UTC
            self.validate_close_end_time = end.astimezone(pytz.utc)

    def validate_month_range(self, output_dir, project_id, start_month, end_month):
        VALIDATION_ERROR = "Validation Error: "
        project_data = self.load_data(project_id)
        errors = []

        for tasks in project_data['tasks']:
            for item in tasks['items']:
                if item['id'] in self.DONT_VALIDATE_THESE_ITEMS:
                    continue
                task_date = datetime.datetime.strptime(item['date'], '%d-%b-%Y').astimezone(pytz.utc)
                error = None
                if task_date < start_month or task_date > end_month:
                    error = "{0} Out Of The Month Range - {1}".format(VALIDATION_ERROR, item)
                if item['duration'] > self.MAX_DURATION:
                    error = "{0} Is longer than the max duration {2} - {1}".format(VALIDATION_ERROR, item, self.MAX_DURATION)
                if error:
                    error = error.encode('utf-8').strip()
                if error:
                    print(error)
                    errors.append(error)

        return errors

    def load_data(self, project_id) -> dict:
        xero_client = self.xero_client

        # Use different times depending on which command
        if self.command == "report-month" or self.command == "report-po":
            time_list = xero_client.time(project_id, self.timesheet_start_time, self.timesheet_end_time)
        else:
            time_list = xero_client.time(project_id, self.validate_close_start_time, self.validate_close_end_time)

        project = xero_client.project(project_id)

        tasks = {}
        for time in time_list:
            user_id = time['userId']
            if user_id in tasks:
                user_time = tasks[user_id]
                user_time.append(self.task_item(time, project_id))
            else:
                user_time = [self.task_item(time, project_id)]
                tasks[user_id] = user_time

        total_hours = 0

        user_tasks = []
        for user_id, user_time_list in tasks.items():
            user = self.user(user_id)
            user_total_duration = 0
            # items is a list of user_time_item
            items = []

            for user_time_item in user_time_list:
                if 'duration' not in user_time_item.keys():
                    raise Exception("User_time_item has an item without a duration attribute. Item:{0} Items:{1}".format(
                        user_time_item, user_time_list))
                task_hours = user_time_item['duration']

                total_hours += task_hours
                user_total_duration += task_hours
                items.append(user_time_item)

            user_tasks.append({
                'userTotalDuration': user_total_duration,
                'userName': user['name'],
                'items': items
            })

        return_dict = {
            'totalHours': total_hours,
            'totalDays': self.round_hours_to_days(total_hours),
            'projectName': project['name'],
            'tasks': user_tasks,
            'startTime': None, # Set below
            'endTime': None # Set below
        }

        if self.command == "report-month" or self.command == "report-po":
            return_dict["startTime"] = self.timesheet_start_time.astimezone(self.local_timezone).strftime("%d %b %Y")
            return_dict["endTime"] = self.timesheet_end_time.astimezone(self.local_timezone).strftime("%d %b %Y")

        else:
            return_dict["startTime"] = self.validate_close_start_time.astimezone(self.local_timezone).strftime("%d %b %Y")
            return_dict["endTime"] = self.validate_close_end_time.astimezone(self.local_timezone).strftime("%d %b %Y")

            # return_dict["startTime"] = (self.start_time + self.SYDNEY_TIME_OFFSET).strftime('%d %b %Y')
            # return_dict["endTime"] = (self.end_time + self.SYDNEY_TIME_OFFSET).strftime('%d %b %Y')

        return return_dict


    def generate_report(self, output_dir, project_ids: list):
        """
        Generates PDF and XLS timesheet

        PO Mode:
        - `project_ids` is a list of relevant project IDs

        Monthly Mode:
        - `project_ids` is a list with 1 element
        - The first element is the project ID that the report should be generated for
        """

        if self.command == "report-month":
            data = self.load_data(project_ids[0])

        elif self.command == "report-po":
            # Merges the data to fit it into the HTML template
            data = {
                "totalHours": 0, # Set below
                "totalDays": None, # Set below
                "projectName": "PO " + str(self.po),
                "tasks": [], # Set below
                "startTime": self.timesheet_start_time.astimezone(self.local_timezone).strftime("%d %b %Y"),
                "endTime": self.timesheet_end_time.astimezone(self.local_timezone).strftime("%d %b %Y")
            }
            for project_id in project_ids:
                print('Generate report by PO, project id=%s' % project_id)
                current_entry = self.load_data(project_id)

                # The variable names are confusing because the template is meant for monthly timesheets, and it's been adapted it to timesheet by PO
                project_tasks = {
                    "userTotalDuration": 0,
                    "userName": current_entry["projectName"],
                    "items": []
                }

                for user_task in current_entry["tasks"]:
                    project_tasks["items"].extend(user_task["items"])
                    for item in user_task["items"]:
                        project_tasks["userTotalDuration"] += item["duration"]

                data["totalHours"] += project_tasks["userTotalDuration"]
                data["tasks"].append(project_tasks)

            data["totalDays"] = self.round_hours_to_days(data["totalHours"])

            # print("Dictionary being sent to template:")
            # pprint(data)

        else:
            raise Exception("Invalid command for generate_report()")
        #pdf
        html = self.generate_html(data)
        self.generate_pdf(html, os.path.join(output_dir, self.report_name(project_ids[0])))

        #xls
        project_name = data['projectName'].replace("/", "_")
        output_file = os.path.join(output_dir, project_name.replace(" ", "_") + '.xls')
        self.generate_xls(data, output_file, project_name)


    def generate_html(self, data):
        template_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'report.html')
        with open(template_file_path) as template_file:
            template = template_file.read()
        return pystache.render(template, data)

    def generate_pdf(self, html, output_file):
        pdfkit.from_string(html, output_file)

    def generate_xls(self, data, output_file, project_name):
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)

        time_list = data

        #Write Header
        workbook_sheet = workbook.create_sheet(title=project_name)
        workbook_sheet.cell(row=1, column=1, value="Detailed Time")
        workbook_sheet.cell(row=2, column=1, value="Darumatic Pty Ltd")
        workbook_sheet.cell(row=3, column=1, value="For the period {} to {}".format(data['startTime'], data['endTime']))
        workbook_sheet.cell(row=6, column=1, value="Date")
        workbook_sheet.cell(row=6, column=2, value="Staff")
        workbook_sheet.cell(row=6, column=3, value="Task")
        workbook_sheet.cell(row=6, column=4, value="Description")
        workbook_sheet.cell(row=6, column=5, value="Total Time")
        workbook_sheet.cell(row=6, column=6, value="Project Name")
        workbook_sheet.cell(row=6, column=7, value="PO")
        workbook_sheet.cell(row=6, column=8, value="Project Owner")
        workbook_sheet.cell(row=4, column=1, value=project_name)

        ROW_OFFSET = 8

        PO = '' if not 'PO' in project_name else project_name.split()[project_name.split().index('PO') + 1]

        # If there's a '-' in the project name, Use the substring after the '-' as the short project name
        # Otherwise, just use itself as the short project name
        short_project_name = project_name if not "-" in project_name else project_name.split("-")[1].strip()
        #short_project_name_with_suffix = "{} - ".format(short_project_name)

        # Defaults to empty string if owner info isn't available
        owner = self.OWNERS[short_project_name] if short_project_name in self.OWNERS.keys() else ""
        if owner == "":
            print(f"Owner info not found for {short_project_name}, leaving it blank")

        i_row = ROW_OFFSET
        for consultant in time_list['tasks']:
            for i, item in enumerate(consultant['items']):
                i_row = i_row + 1
                workbook_sheet.cell(row=i_row, column=1, value=item['date'])
                workbook_sheet.cell(row=i_row, column=2, value=item['userName'])
                workbook_sheet.cell(row=i_row, column=3, value=item['taskName'])
                workbook_sheet.cell(row=i_row, column=4, value=item['taskDescription'])
                workbook_sheet.cell(row=i_row, column=5, value=item['duration'])
                workbook_sheet.cell(row=i_row, column=6, value=short_project_name)
                workbook_sheet.cell(row=i_row, column=7, value=PO)
                workbook_sheet.cell(row=i_row, column=8, value=owner)

        workbook.save(output_file)
        print("Saving workbook in {}".format(output_file))

    def report_name(self, project_id):
        if self.command == "report-month":
            project = self.xero_client.project(project_id)
            project_name = project['name']
            report_name = ''
            skip = False
            for c in project_name:
                if c.isalpha() or c.isdigit():
                    report_name += c.lower()
                    skip = False
                else:
                    if not skip:
                        report_name += '_'
                        skip = True

            report_name += '_'
            report_name += self.timesheet_start_time.astimezone(self.local_timezone).strftime('%d-%b-%Y').lower()
            report_name += '-'
            report_name += self.timesheet_end_time.astimezone(self.local_timezone).strftime('%d-%b-%Y').lower()
            report_name += '.' + ''.join(random.sample('0123456789', 3))
            report_name += '.pdf'

        elif self.command == "report-po":
            # Generate timesheet by PO case
            report_name = f"po_{str(self.po)}_timesheet_ending_{self.timesheet_end_time.astimezone(self.local_timezone).strftime('%d-%b-%Y')}.pdf"

        return report_name

    def task_item(self, time, project_id):
        user = self.user(time['userId'])
        task = self.task(time['taskId'], project_id)

        # Interpretate UTC string and then save it as local time
        local_time = dateutil.parser.parse(time['dateUtc']).astimezone(self.local_timezone)
        # local_time = datetime.datetime.strptime(time['dateUtc'], '%Y-%m-%dT%H:%M:%SZ').astimezone(self.local_timezone)
        ret = {
            'userName': user['name'],
            'taskName': task['name'],
            'id': task['taskId'],
            'date': local_time.strftime('%d-%b-%Y'),
            'duration': self.round_minutes_to_hours(time['duration'])
        }

        try:
            description = time['description']
        except Exception as e:
            print("Error capturing description from:{0} Falling back to empty string".format(str(time)))
            description = ' '

        ret['taskDescription'] = description

        return ret

    def round_minutes_to_hours(self, minutes):
        return round(float(minutes) / 60, 4)

    def round_hours_to_days(self, hours):
        return round(float(hours) / 8, 2)

    def user(self, user_id):
        xero_client = self.xero_client
        if user_id in self.cache_users:
            return self.cache_users[user_id]
        else:
            user = xero_client.user(user_id)
            self.cache_users[user_id] = user
            return user

    def task(self, task_id, project_id):
        xero_client = self.xero_client
        if task_id in self.cache_tasks:
            return self.cache_tasks[task_id]
        else:
            task = xero_client.task(project_id, task_id)
            self.cache_tasks[task_id] = task
            return task

    def get_all_projects(self):
        return self.xero_client.get_items('https://api.xero.com/projects.xro/2.0/projects',
                                          one_page=False)

    def get_active_projects(self):
        return self.xero_client.get_items('https://api.xero.com/projects.xro/2.0/projects?states=INPROGRESS',
                                          one_page=False)

    def print_active_projects(self):
        for items in self.get_active_projects():
            for item in items['items']:
                print("Name: {0}, ProjectID: {1}, Status: {2}, ContactId: {3}".format(item["name"], item["projectId"],
                                                                                      item["status"],
                                                                                      item["contactId"]))

    def create_monthly_time_sheets(self):
        if os.path.exists(self.output):
            shutil.rmtree(self.output)
            os.makedirs(self.output)
        else:
            os.makedirs(self.output)

        if self.command == "report-month":
            # This should be last month unless otherwise specified
            # E.g. when run in January 2021, target_project is "202012"
            # E.g. when run in March 2021, target_project is "202102"
            # This can be changed with $PROJECT_NAME
            target_project = self.target_month_string

        elif self.command == "report-po":
            # For generating timesheet by po
            if not self.po:
                raise Exception("Can't generate report by PO, no PO given. Please setup the PO environment variable")

            target_project = "PO " + str(self.po)

        else:
            raise Exception(f"{self.command} isn't a valid command for generating timesheets")

        print(f"Creating timesheet for projects named {target_project}")

        # Create active projects json file
        os.mkdir(self.output+"/backup/")
        f = open(self.output+"/backup/all_projects.json", "w+")
        json.dump(self.get_active_projects(), f)
        f.close()

        # Flag to check whether there was a project with name target_project
        project_found = False
        # For report generation by PO
        projects_ids_with_target_po = []

        for items in self.get_all_projects():
            for item in items['items']:
                # Only generate timesheet for projects with target_project in it
                if target_project not in item['name']:
                    continue

                # Replaces spaces and slashes with underscores
                item['name'] = item['name'].replace(" ", "_").replace("/", "_")

                project_found = True
                print('Generate Xero report for project %s between %s %s to %s' % (
                    item["projectId"], self.timesheet_start_time, self.timesheet_end_time, self.output))
                print("Name: {0}, ProjectID: {1}, Status: {2}, ContactId: {3}".format(item["name"],
                                                                                      item["projectId"],
                                                                                      item["status"],
                                                                                      item["contactId"]))
                self.backup_data(item["projectId"], item["name"])

                self.generate_report(self.output, [item["projectId"]])
                #if self.command == "report-month":
                #    # Run generate report here when using report-month
                #    self.generate_report(self.output, [item["projectId"]])
                #else:
                #    projects_ids_with_target_po.append(item["projectId"])

        if not project_found:
            raise Exception(f"No project named {target_project} was found on Xero!")

        # if self.command == "report-po":
        #     # Run generate report here when using report-po
        #     self.generate_report(self.output, projects_ids_with_target_po)

        if self.suppress_email:
            print("Timesheet not sent because suppress flag is true")

        else:
            # Sends the generated reports as an email
            attachment_paths = []
            attachments = os.listdir(self.output)
            # Gets all files in the output folder that end with .xls or .pdf
            for attachment in attachments:
                full_path = os.path.join(self.output, attachment)
                if attachment.endswith((".xls", ".pdf")) and not os.path.isdir(full_path):
                    attachment_paths.append(full_path)

            if attachment_paths == []:
                raise Exception("Couldn't find any files to attach to email")

            # Changes the email title
            if self.command == "report-month":
                XeroEmailSender.send_timesheet(attachment_paths,    self.timesheet_start_time.astimezone(self.local_timezone), "month")

            else:
                XeroEmailSender.send_timesheet(attachment_paths, self.timesheet_end_time.astimezone(self.local_timezone), "po", po=self.po)


    def validate_projects(self, reporter):
        # Assume we are working in UTC time
        month_start = self.validate_close_start_time
        month_end = self.validate_close_end_time
        DARUMATIC_INIT_TIME = '2017-02-20'
        start_validation_time = DARUMATIC_INIT_TIME
        self.add_project_times(start_validation_time, None)
        end_validation_time = self.validate_close_start_time + datetime.timedelta(days=5000)
        self.add_project_times(None, end_validation_time.strftime('%Y-%m-%d'))
        start_validation_time, end_validation_time = self.validate_close_start_time, self.validate_close_end_time

        print('Validating Active Projects..')
        print("Start Validation Time: {0} \n"
              "End Validation time: {1} \n"
              "Start Month: {2} \n"
              "End Month: {3} \n"
              "Start Time: {4} \n"
              "End Time: {5}".format(
            start_validation_time.astimezone(self.local_timezone),
            end_validation_time.astimezone(self.local_timezone),
            month_start.astimezone(self.local_timezone),
            month_end.astimezone(self.local_timezone),
            self.validate_close_start_time.astimezone(self.local_timezone),
            self.validate_close_end_time.astimezone(self.local_timezone)
        ))
        errors = {}
        amount_of_errors = 0

        for items in self.get_all_projects():
            pprint(items)
            for item in items['items']:
                # Check active projects timestamp in project name'
                if self.filter not in item['name']:
                    continue
                print("Name: {0}, ProjectID: {1}, Status: {2}, ContactId: {3}".format(item["name"], item["projectId"],
                                                                                      item["status"],
                                                                                      item["contactId"]))
                errors[item["name"]] = []
                # validate that projects have timestamps------------------
                timestamp_error = self.validate_timestamp(item['name'])
                if timestamp_error:
                    timestamp_error = timestamp_error.encode('utf-8').strip()
                    errors[item["name"]].append(timestamp_error)
                    amount_of_errors += 1
                # validate tasks month range ------------------
                project_errors = reporter.validate_month_range(self.output, item["projectId"], month_start, month_end)
                if len(project_errors) > 0:
                    errors[item["name"]] = project_errors
                    amount_of_errors = len(project_errors) + amount_of_errors

        #validate that employees worked all working days-----------------
        # Working days for previous month
        # year = self.filter[0:4]
        # month = self.filter[4:6]
        # last_month_workdays = self.generate_work_days(year, month, self.filter)
        # for employee in self.EMPLOYEES:
        #     errors[employee["userName"]] = []
        #     last_month_employee_workdays = []
        #     for workdays in employee["working_days"]:
        #         if (year + "-" + month) in workdays["date"]:
        #             last_month_employee_workdays.append(workdays)
        #     # Validate
        #     # @days are required work days
        #     # @workdays are the days when employees attend
        #     # @workdays are dictionaries {"date":"" , "duration":""}
        #     for days in last_month_workdays:
        #         attend = False
        #         for workdays in last_month_employee_workdays:
        #             if days == workdays["date"]:
        #                 attend = True
        #                 if workdays["duration"] < 8:
        #                     error = "Working hour error: {0} attend less than 8 hours on {1}".format(employee["userName"], days)
        #                     errors[employee["userName"]].append(error)
        #
        #         if not attend:
        #             error = "Working hour error: {0} didn't attend on {1}".format(employee["userName"], days)
        #             errors[employee["userName"]].append(error)
        #         # Update errors amount
        #     amount_of_errors = len(errors[employee["userName"]]) + amount_of_errors

        print("*" * 80)
        print("List of Validation errors")
        pprint(errors)
        print("*" * 80)
        if amount_of_errors == 0:
            result = "There are no errors in {0}. All tasks are validated successfully!!".format(self.filter)
            print(result)
        else:
            result = "There were a total of {0} errors in {1}.".format(amount_of_errors, self.filter)
            print(result)
            if not self.suppress_email:
                XeroEmailSender.send_validation_report(result, datetime.datetime.now(self.local_timezone), errors)
            else:
                print("Validation report not sent because suppress flag is true")
        return amount_of_errors

    def backup_data(self, project_id, project_name):
        # Create current project's folder
        project_folder = self.output + "/backup/" + project_name
        os.mkdir(project_folder)
        # Create time entry file
        xero_client = self.xero_client
        time_list = xero_client.time(project_id, self.timesheet_start_time, self.timesheet_end_time)
        time_list_file = open(project_folder+"/time_entries.json", "w+")
        json.dump(time_list, time_list_file)
        time_list_file.close()
        # Create tasks file
        task_list = xero_client.get_tasks(project_id)
        task_list_file = open(project_folder+"/tasks.json", "w+")
        json.dump(task_list, task_list_file)
        task_list_file.close()
        # Create users file
        user_list = []
        for time_entry in time_list:
            user_id = time_entry["userId"]
            user = xero_client.user(user_id)
            user_list.append(user)
        user_list_file = open(project_folder+"/users.json", "w+")
        json.dump(user_list, user_list_file)
        user_list_file.close()

    def close_previous_month_projects(self):
        counter = 0
        for items in self.get_all_projects():
            for item in items['items']:
                if self.filter not in item['name'] and item['status'] == "INPROGRESS":
                    data = {"status": "CLOSED"}
                    project_id = item["projectId"]
                    response = self.xero_client.patch_projects(
                        'https://api.xero.com/projects.xro/2.0/Projects/' + project_id, data)
                    if response == 204:
                        print("Closing project: " + item["name"] + "  SUCCEED!")
                        counter += 1
                    else:
                        print("Closing project: " + item["name"] + " FAILED!")
            if counter == 0:
                print("No projects have been closed")
            else:
                print("{0} projects have been closed".format(counter))

    def validate_timestamp(self, project_name):
        VALIDATION_ERROR = "Timestamp Validation Error:"
        try:
            index = project_name.index("-")
            timestamp = project_name[:index].strip()
        except ValueError:
            return "{0} Doesn't contain timestamp starting with '-'.".format(VALIDATION_ERROR)

        # Length
        if len(timestamp) != 6:
            return "{0} Cannot find timestamp.".format(VALIDATION_ERROR)

        # Non-numeric character
        if not timestamp.isdigit():
            return "{0} Timestamp contains illegal characters, only numbers allowed.".format(VALIDATION_ERROR)

        # Year
        #todo: change this for a isdate validation
        if timestamp[0:2] != "20":
            return "{0} Invalid year. Year should be : 20XX".format(VALIDATION_ERROR)

        #  Month
        if not (1 <= int(timestamp[4:]) <= 12):
            return "{0} Invalid month. Month should from [01 - 12].".format(VALIDATION_ERROR)

        return None

    def generate_work_days(self, year, month, current_filter):
        # Get previous month
        year = current_filter[0:4]
        month = current_filter[4:6]
        if month == "01":
            year = str(int(year)-1)
            month = "12"
        else:
            if int(month) < 10:
                month = "0" + str(int(month) - 1)
            else:
                month = str(int(month) - 1)

        workdays = []
        for week in calendar.monthcalendar(int(year), int(month)):
            del week[5:7]
            while 0 in week:
                week.remove(0)
            for day in week:
                if day < 10:
                    date = year + "-" + month + "-" + "0" + str(day)
                else:
                    date = year + "-" + month + "-" + str(day)
                workdays.append(date)

        # Remove holidays
        for year in self.NSW_HOLIDAYS:
            if year["year"] == "2020":
                for holidays in year["holidays"]:
                    if "2020-04" in holidays and holidays in workdays:
                        workdays.remove(holidays)

        return workdays


    def _get_time_sheet_project_name(self) -> str:
        """
        Gets the project name for the timesheet
        - Defaults to last month if `$PROJECT_NAME` isn't found
        """
        if "PROJECT_NAME" in os.environ:
            print(f"Environmental variable PROJECT_NAME found, using {os.getenv('PROJECT_NAME')}")
            return os.getenv("PROJECT_NAME")

        now = datetime.datetime.now(self.local_timezone)
        year = now.year
        month = now.month - 1
        if month == 0:
            month = 12
            year -= 1

        print(f"Environmental variable PROJECT_NAME not found, defaulting to {year}{month:02d}")
        return f"{year}{month:02d}"


    def _get_time_sheet_start_time(self, command_line_time: str) -> datetime.datetime:
        """
        Sets the time ranges for generating report

        `report-month`
        - Uses `self.target_month_string`
        - Returns the first day of that month

        `report-po`
        - The date ranges for `report-po` were given via command line arguments
        - Parse the date given by the command line
        - If no start time was given via command line arguments, it returns `DARUMATIC_START_TIME`
        """
        if self.command == "report-month":
            target_month = datetime.datetime.strptime(self.target_month_string, "%Y%m")
            # This datetime object is still in Sydney time, but is timezone aware
            localised_datetime_object = self.local_timezone.localize(target_month)

            # Turn it back into UTC for compatability with existing code
            return localised_datetime_object.astimezone(pytz.utc)

        elif self.command == "report-po":
            if command_line_time:
                # Start time was given via the command line
                # Parse the string
                start = datetime.datetime.strptime(command_line_time, "%Y-%m-%d")
                # Localise it with the self.local_timezone
                start = self.local_timezone.localize(start)
                # Convert it back to UTC for compatability
                start = start.astimezone(pytz.utc)

            else:
                # Defaults to Darumatic start time if doing report-po and no start time was provided
                start = DARUMATIC_START_TIME
                print("Running report-po but no start time provided, defaulting to Darumatic start time")

            return start



    def _get_time_sheet_end_time(self, command_line_time: str) -> datetime.datetime:
        """
        Same as `self._get_time_sheet_start_time` but for end time

        When running `report-po` and no end time was given via command line arguments, it returns now
        """
        if self.command == "report-month":
            target_month = datetime.datetime.strptime(self.target_month_string, "%Y%m")
            month_last_day = calendar.monthrange(target_month.year, target_month.month)[1]
            # This datetime object is still in Sydney time, but is timezone aware
            # Day is set to last day of the month
            localised_datetime_object = self.local_timezone.localize(target_month.replace(day=month_last_day, hour=23, minute=59, second=59))

            # Turn it back into UTC for compatability with existing code
            return localised_datetime_object.astimezone(pytz.utc)

        elif self.command == "report-po":
            if command_line_time:
                # End time was given via the command line
                # Parse the string
                end = datetime.datetime.strptime(command_line_time, "%Y-%m-%d")
                # Localise it with self.local_timezone
                end = self.local_timezone.localize(end)
                # Convert it back to UTC for compatability
                end = end.astimezone(pytz.utc)

            else:
                # Defaults to now if doing report-po and no end time was provided
                # Localises utcnow with the local timezone and converts it back to UTC for compatibility
                end = self.local_timezone.localize(datetime.datetime.utcnow()).astimezone(pytz.utc)
                print("Running report-po but no end time provided, defaulting to now")

            return end



    def _get_local_time_zone(self):
        """
        Gets the timezone from `$LOCAL_TIMEZONE`
        - Defaults to `DEFAULT_TIMEZONE`
        """
        if "LOCAL_TIMEZONE" in os.environ:
            timezone = os.getenv("LOCAL_TIMEZONE")
            print(f"Setting timezone as {timezone}")
            try:
                timezone = pytz.timezone(timezone)
            except pytz.UnknownTimeZoneError:
                raise Exception(f"Couldn't find timezone {timezone}, use the 'TZ database name' field from https://en.wikipedia.org/wiki/List_of_tz_database_time_zones")
        else:
            timezone = DEFAULT_TIMEZONE

        return timezone

    def _email_env_var_present(self) -> bool:
        """
        Checks whether the environmental variables for emailing are all present:
        - `$SENDER`
        - `$RECEIVER`
        - `$SENDGRID_API_KEY`
        """
        return "SENDER" in os.environ and "RECEIVER" in os.environ and "SENDGRID_API_KEY" in os.environ

if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.realpath(__file__))

    parser = argparse.ArgumentParser("Generate Xero project reports")
    parser.add_argument('--client-id', type=str, required=True)
    parser.add_argument('--client-secret', type=str, required=True)
    parser.add_argument('--refresh-token', type=str, required=True)
    parser.add_argument('--tenant-id', type=str, required=True)
    parser.add_argument('--start-time', type=str, required=False, default=None)
    parser.add_argument('--end-time', type=str, required=False, default=None)
    parser.add_argument('--output', type=str, default=os.path.join(current_dir, "out"))
    parser.add_argument('--po', type=int, required=False, default=None)

    # Optional argument to skip loading in owners data
    parser.add_argument('--skip-owners', type=bool, default=False)
    # Optional argument to not send emails
    parser.add_argument('--suppress-email', type=bool, default=False)


    command = sys.argv[1]
    args = parser.parse_args(sys.argv[2:])
    reporter = XeroReport(args, command)

    if command == "report-month":
        reporter.create_monthly_time_sheets()
    elif command == "report-po":
        reporter.create_monthly_time_sheets()
    elif command == "close":
        reporter.close_previous_month_projects()
    elif command == "validate":
        amount_of_errors = reporter.validate_projects(reporter)
        print("The Validate function informed that there were {0} errors. \n"
              "Please check the logs above for more information.".format(amount_of_errors))
    else:
        print("Invalid command")
        sys.exit(2)

    print("Xero-Automation - Finished successfully!")
    sys.exit(0)
